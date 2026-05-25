"""
excel_logger.py
---------------
Appends every signal scan to data/signals.xlsx for permanent history.

Each row = one signal from one scan, timestamped. This builds the
dataset you'll use to forward-test the strategy (review how past
signals actually performed).

Creates the file with headers on first run; appends thereafter.
"""

import logging
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from src.config import OUTPUT, PROJECT_ROOT

logger = logging.getLogger(__name__)

# Columns logged, in order
COLUMNS = [
    "scan_date", "scan_time", "ticker", "tier", "category",
    "short_float_pct", "short_ratio", "current_price", "high_100d",
    "sma_50", "distance_from_high_pct", "is_breakout",
    "volume_confirmed", "high_conviction",
]


def _excel_path() -> Path:
    """Resolve the signals.xlsx path from settings, relative to project root."""
    rel = OUTPUT["excel_file"]  # "data/signals.xlsx"
    return PROJECT_ROOT / rel


def log_signals(df) -> int:
    """
    Append the current scan's signals to the Excel log.

    Args:
        df: DataFrame of qualifying signals (post-classification).

    Returns:
        Number of rows appended.
    """
    if df is None or df.empty:
        logger.info("No signals to log to Excel.")
        return 0

    path = _excel_path()
    path.parent.mkdir(parents=True, exist_ok=True)

    now = datetime.now()
    scan_date = now.strftime("%Y-%m-%d")
    scan_time = now.strftime("%H:%M:%S")

    # Build the rows to append
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "scan_date": scan_date,
            "scan_time": scan_time,
            "ticker": r.get("ticker", ""),
            "tier": r.get("Tier", ""),
            "category": r.get("category", ""),
            "short_float_pct": r.get("short_float_pct", ""),
            "short_ratio": r.get("Short Ratio", ""),
            "current_price": r.get("current_price", ""),
            "high_100d": r.get("high_100d", ""),
            "sma_50": r.get("sma_50", ""),
            "distance_from_high_pct": r.get("distance_from_high_pct", ""),
            "is_breakout": bool(r.get("is_breakout", False)),
            "volume_confirmed": bool(r.get("volume_confirmed", False)),
            "high_conviction": bool(r.get("high_conviction", False)),
        })

    new_df = pd.DataFrame(rows, columns=COLUMNS)

    if path.exists():
        # Append to existing file
        existing = pd.read_excel(path)
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df

    # Write out
    combined.to_excel(path, index=False, sheet_name="Signals")

    # Bold the header row
    try:
        wb = load_workbook(path)
        ws = wb["Signals"]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Reasonable column widths
        widths = {"A": 12, "B": 10, "C": 8, "D": 10, "E": 14,
                  "F": 14, "G": 11, "H": 13, "I": 11, "J": 9,
                  "K": 20, "L": 11, "M": 16, "N": 15}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        wb.save(path)
    except Exception as e:
        logger.warning(f"Could not apply Excel formatting: {e}")

    logger.info(f"Logged {len(new_df)} signal(s) to {path.name} (total rows: {len(combined)}).")
    return len(new_df)